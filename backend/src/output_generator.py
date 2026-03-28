"""
output_generator.py - Generate JSON, CSV, and Excel output
"""

import json
import csv
from typing import Dict, Any
from pathlib import Path
from datetime import datetime

from src.utils.logger import get_logger

logger = get_logger(__name__)


class OutputGenerator:
    def __init__(self, output_dir: str = 'output'):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        self.logger = get_logger(__name__)

    def save_json(self, data: Dict[str, Any], filename: str = 'extracted_data.json') -> bool:
        try:
            with open(self.output_dir / filename, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
            return True
        except Exception as e:
            self.logger.error(f"JSON error: {e}")
            return False

    def save_csv(self, data: Dict[str, Any], filename: str = 'extracted_data.csv') -> bool:
        try:
            flat = self._flatten_dict(data)
            with open(self.output_dir / filename, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=flat.keys())
                writer.writeheader()
                writer.writerow(flat)
            return True
        except Exception as e:
            self.logger.error(f"CSV error: {e}")
            return False

    def save_xlsx(self, data: Dict[str, Any], filename: str = 'extracted_data.xlsx', source_filename: str = '') -> bool:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import (
                Font, PatternFill, Alignment, Border, Side, numbers
            )

            wb = Workbook()
            ws = wb.active
            ws.title = "Extracted Invoice Data"

            # ── Colors ──
            GREEN       = "217346"
            GREEN_LIGHT = "E8F5EE"
            GREEN_MID   = "C6E8D4"
            HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
            LABEL_FONT  = Font(name="Arial", bold=True, color="185A36", size=10)
            VALUE_FONT  = Font(name="Arial", size=10)
            MONO_FONT   = Font(name="Courier New", size=10)
            GRAY_FONT   = Font(name="Arial", size=9, color="888888")

            green_fill  = PatternFill("solid", fgColor=GREEN)
            label_fill  = PatternFill("solid", fgColor=GREEN_LIGHT)
            alt_fill    = PatternFill("solid", fgColor="F9F9F9")
            mid_fill    = PatternFill("solid", fgColor=GREEN_MID)

            thin = Side(style="thin", color="CCCCCC")
            med  = Side(style="medium", color=GREEN)
            thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
            med_left    = Border(left=Side(style="medium", color=GREEN), right=thin, top=thin, bottom=thin)

            center = Alignment(horizontal="center", vertical="center")
            left   = Alignment(horizontal="left",   vertical="center")
            right  = Alignment(horizontal="right",  vertical="center")

            # ── Title Row ──
            ws.merge_cells("A1:E1")
            ws["A1"] = "📄 Invoice Data Extraction Report"
            ws["A1"].font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
            ws["A1"].fill = green_fill
            ws["A1"].alignment = center
            ws.row_dimensions[1].height = 32

            # ── Subtitle ──
            ws.merge_cells("A2:E2")
            ws["A2"] = f"Source File: {source_filename or 'Invoice'}   |   Extracted: {datetime.now().strftime('%d %b %Y, %H:%M')}"
            ws["A2"].font = GRAY_FONT
            ws["A2"].fill = PatternFill("solid", fgColor="F0F7F3")
            ws["A2"].alignment = center
            ws.row_dimensions[2].height = 18

            ws.row_dimensions[3].height = 8  # spacer

            # ── Column Headers ──
            headers = ["#", "Field", "Extracted Value", "Status", "Notes"]
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=4, column=col, value=h)
                c.font = HEADER_FONT
                c.fill = green_fill
                c.alignment = center
                c.border = thin_border
            ws.row_dimensions[4].height = 22

            # ── Data Rows ──
            fields = [
                ("Invoice Number",  data.get("invoice_number", ""),           "Unique invoice identifier"),
                ("Invoice Date",    data.get("invoice_date", ""),             "Date of issue"),
                ("Vendor Name",     (data.get("vendor") or {}).get("name",""),"Supplier / From"),
                ("Customer Name",   (data.get("customer") or {}).get("name",""),"Bill To / Buyer"),
                ("Subtotal",        data.get("subtotal", 0),                  "Amount before tax"),
                ("Tax Amount",      data.get("tax", 0),                       "GST / VAT / Tax"),
                ("Total Amount",    data.get("total", 0),                     "Grand total payable"),
                ("Validation",      "PASSED" if data.get("validation_passed") else "FAILED",
                                    "; ".join(data.get("validation_errors", []) or ["All checks passed"])),
            ]

            currency_rows = {5, 6, 7}  # subtotal, tax, total (1-indexed from fields)

            for i, (field, value, note) in enumerate(fields):
                row = i + 5
                is_alt = i % 2 == 1
                row_fill = alt_fill if is_alt else PatternFill("solid", fgColor="FFFFFF")

                # # column
                n = ws.cell(row=row, column=1, value=i + 1)
                n.font = Font(name="Arial", size=9, color="999999")
                n.fill = label_fill
                n.alignment = center
                n.border = thin_border

                # Field label
                lbl = ws.cell(row=row, column=2, value=field)
                lbl.font = LABEL_FONT
                lbl.fill = label_fill
                lbl.alignment = left
                lbl.border = med_left

                # Value
                is_currency = (i + 1) in currency_rows
                val_cell = ws.cell(row=row, column=3)
                if is_currency:
                    val_cell.value = float(value) if value else 0.0
                    val_cell.number_format = '₹#,##0.00'
                    val_cell.font = Font(name="Courier New", bold=True, size=10,
                                        color="1A5C2E" if float(value or 0) > 0 else "999999")
                    val_cell.alignment = right
                else:
                    val_cell.value = str(value) if value else ""
                    val_cell.font = MONO_FONT if value else Font(name="Arial", size=10, color="CCCCCC", italic=True)
                    if not value:
                        val_cell.value = "(not detected)"
                    val_cell.alignment = left
                val_cell.fill = row_fill
                val_cell.border = thin_border

                # Status
                stat_cell = ws.cell(row=row, column=4)
                if field == "Validation":
                    stat_cell.value = "✅ PASSED" if value == "PASSED" else "❌ FAILED"
                    stat_cell.font = Font(name="Arial", bold=True, size=10,
                                         color="1A7A3E" if value == "PASSED" else "C0392B")
                    stat_cell.fill = PatternFill("solid", fgColor="E8F5EE" if value == "PASSED" else "FDECEA")
                elif is_currency:
                    stat_cell.value = "✓ Found" if float(value or 0) > 0 else "⚠ Missing"
                    stat_cell.font = Font(name="Arial", size=10,
                                         color="1A7A3E" if float(value or 0) > 0 else "E67E22")
                    stat_cell.fill = row_fill
                else:
                    stat_cell.value = "✓ Found" if value else "⚠ Missing"
                    stat_cell.font = Font(name="Arial", size=10,
                                         color="1A7A3E" if value else "E67E22")
                    stat_cell.fill = row_fill
                stat_cell.alignment = center
                stat_cell.border = thin_border

                # Notes
                note_cell = ws.cell(row=row, column=5, value=note)
                note_cell.font = Font(name="Arial", size=9, color="888888", italic=True)
                note_cell.fill = row_fill
                note_cell.alignment = left
                note_cell.border = thin_border

                ws.row_dimensions[row].height = 20

            # ── Summary box ──
            sum_row = len(fields) + 6
            ws.row_dimensions[sum_row] = ws.row_dimensions[sum_row]
            ws.row_dimensions[sum_row].height = 8  # spacer

            ws.merge_cells(f"B{sum_row+1}:C{sum_row+1}")
            ws[f"B{sum_row+1}"] = "GRAND TOTAL"
            ws[f"B{sum_row+1}"].font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
            ws[f"B{sum_row+1}"].fill = green_fill
            ws[f"B{sum_row+1}"].alignment = left
            ws[f"B{sum_row+1}"].border = thin_border

            total_val = data.get("total", 0)
            ws[f"D{sum_row+1}"] = float(total_val) if total_val else 0.0
            ws[f"D{sum_row+1}"].number_format = '₹#,##0.00'
            ws[f"D{sum_row+1}"].font = Font(name="Courier New", bold=True, size=12, color="FFFFFF")
            ws[f"D{sum_row+1}"].fill = green_fill
            ws[f"D{sum_row+1}"].alignment = right
            ws[f"D{sum_row+1}"].border = thin_border
            ws.row_dimensions[sum_row+1].height = 24

            # ── Column widths ──
            ws.column_dimensions["A"].width = 5
            ws.column_dimensions["B"].width = 22
            ws.column_dimensions["C"].width = 30
            ws.column_dimensions["D"].width = 16
            ws.column_dimensions["E"].width = 38

            # ── Freeze header ──
            ws.freeze_panes = "A5"

            filepath = self.output_dir / filename
            wb.save(str(filepath))
            self.logger.info(f"XLSX saved: {filepath}")
            return True

        except Exception as e:
            self.logger.error(f"XLSX error: {e}")
            import traceback; traceback.print_exc()
            return False

    @staticmethod
    def _flatten_dict(data: Dict[str, Any], parent_key: str = '') -> Dict[str, Any]:
        items = []
        for k, v in data.items():
            new_key = f"{parent_key}_{k}" if parent_key else k
            if isinstance(v, dict):
                items.extend(OutputGenerator._flatten_dict(v, new_key).items())
            elif isinstance(v, list):
                items.append((new_key, str(v)))
            else:
                items.append((new_key, v))
        return dict(items)

    def generate_all_outputs(self, data: Dict[str, Any], base_name: str = 'invoice', source_filename: str = '') -> bool:
        try:
            self.save_json(data, f'{base_name}.json')
            self.save_csv(data, f'{base_name}.csv')
            self.save_xlsx(data, f'{base_name}.xlsx', source_filename=source_filename)
            self.logger.info(f"All outputs generated for {base_name}")
            return True
        except Exception as e:
            self.logger.error(f"Output error: {e}")
            return False


def save_invoice_data(data: Dict[str, Any], output_dir: str = 'output') -> bool:
    return OutputGenerator(output_dir).generate_all_outputs(data)
